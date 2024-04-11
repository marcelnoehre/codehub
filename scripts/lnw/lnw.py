import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font


def main():
    start = datetime.strptime(input('Start:'), "%d.%m.%Y")
    end = datetime.strptime(input('End:'), "%d.%m.%Y")
    shortage = input('Shortage:')
    sprint = input('Sprint:')

    create_lnw(extract_time(start, end, shortage, sprint), shortage, sprint)


def extract_time(start, end, shortage, sprint):
    months = []
    lnw_mde2 = []
    lnw_SCB = []
    daily = 0

    dates = [
        [str(start.year - 1 if start.month - 1 == 0 else start.year), str(12 if start.month - 1 == 0 else start.month - 1).zfill(2)],
        [str(start.year), str(start.month).zfill(2)],
        [str(start.year + 1 if start.month + 1 == 13 else start.year), str(1 if start.month + 1 == 13 else start.month + 1).zfill(2)],
    ]

    if end.month > start.month or end.year > start.year:
        if(str(1 if end.month + 1 == 13 else end.month).zfill(2)) not in months:
            months.append([str(end.year + 1 if end.month + 1 == 13 else end.year), str(1 if end.month + 1 == 13 else end.month).zfill(2)])
    
    for date in dates:
        try:
            df = pd.read_excel(f'{date[0]}_{date[1]}_{shortage}_Tagesberichte.xlsx')
            for index, row in df.iterrows():
                try:
                    if start <= datetime.strptime(str(row.iloc[0]).split(', ')[1], "%d.%m.%Y") <= end:
                        if row.iloc[1] == 'MDE2.0':
                            if row.iloc[5] == 'Scrum-Daily':
                                daily += 0.25
                            else:
                                lnw_mde2.append([row.iloc[3], 'Kettler St. Ingbert', 'MDE2.0', row.iloc[0].split(', ')[1], row.iloc[6], row.iloc[4], f'Sprint-{sprint}', '', '', '' , f'{shortage}: {row.iloc[5]}'])
                        elif row.iloc[1] == 'SCB':
                            lnw_SCB.append([row.iloc[3], 'Schmitz Cargobull', 'SCB', row.iloc[0].split(', ')[1], row.iloc[6], row.iloc[4], '200435', '', '', '' , f'{shortage}: {row.iloc[5]}' ])
                except Exception:
                    pass
        except Exception:
            pass

    lnw_mde2.append(['P000001576', 'Kettler St. Ingbert', 'MDE2.0', f'{str(end.day).zfill(2)}.{str(end.month).zfill(2)}.{end.year}', daily, 'M', f'Sprint-{sprint}', '', '', '' , f'{shortage}: Projektplanung'])
    
    return [lnw_mde2, lnw_SCB]


def create_lnw(lnw, shortage, sprint):
    workbook = load_workbook('template.xlsx')
    sheet = workbook.active
    sheet.title = f'Sprint-{sprint}'

    for i in range(len(lnw)):
        for j in range(len(lnw[i])):

            cell = sheet.cell(row = i + 2, column = j + 1, value = lnw[i][j])
            cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))

            if sheet.cell(row = 1, column = j + 1).value == 'STD':
                cell.alignment = Alignment(horizontal='center', vertical='center')
                decimal = len(str(cell.value).split('.')[1]) if '.' in str(cell.value) else 0
                cell.number_format = '0' if decimal == 0 else '0.' + '0' * decimal

            elif sheet.cell(row = 1, column = j + 1).value == 'AP und Tätigkeit':
                cell.alignment = Alignment(horizontal='left', vertical='center')

            elif sheet.cell(row = 1, column = j + 1).value == 'Projektnummer':
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)

            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    workbook.save(f'Leistungsnachweise-{shortage}-Sprint {sprint}.xlsx')


if __name__ == '__main__':
    main()
